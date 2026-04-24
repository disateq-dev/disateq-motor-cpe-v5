"""
generar_excel_prueba.py
Genera archivo Excel simulado DisateQ POS™ v1.2
"""

from openpyxl import Workbook
from datetime import datetime, timedelta
import random

def generar_excel_prueba():
    wb = Workbook()
    ws = wb.active
    ws.title = "_CPE"
    
    # HEADERS (Fila 1) - Contrato POS v1.2
    headers = [
        'cpe_tipo', 'cpe_serie', 'cpe_numero', 'cpe_fecha', 'cpe_moneda',
        'cli_tipo_doc', 'cli_nro_doc', 'cli_nombre', 'cli_direccion',
        'item_codigo', 'item_descripcion', 'item_cantidad', 'item_unidad',
        'item_precio_unitario', 'item_valor_unitario',
        'item_subtotal_sin_igv', 'item_igv', 'item_total',
        'item_afectacion_igv', 'item_unspsc',
        'venta_subtotal', 'venta_exonerada', 'venta_inafecta', 'venta_icbper',
        'venta_igv', 'venta_total',
        'pago_forma', 'pago_monto',
        'estado'
    ]
    ws.append(headers)
    
    # DATOS SIMULADOS
    comprobantes = [
        # Boleta 1 - Cliente VARIOS (2 items)
        {
            'tipo': '03', 'serie': 'B001', 'numero': 1, 
            'cliente': ('1', '00000000', 'CLIENTES VARIOS', 'AV. PRINCIPAL 123'),
            'items': [
                ('P001', 'Arroz Costeño 1kg', 2, 'NIU', 5.00),
                ('P002', 'Azúcar Rubia 1kg', 3, 'NIU', 3.50)
            ]
        },
        # Factura 1 - Empresa (1 item)
        {
            'tipo': '01', 'serie': 'F001', 'numero': 1,
            'cliente': ('6', '20123456789', 'EMPRESA DEMO S.A.C.', 'AV. LARCO 456'),
            'items': [
                ('P003', 'Aceite Primor 1L', 10, 'NIU', 12.00)
            ]
        },
        # Boleta 2 - DNI (3 items)
        {
            'tipo': '03', 'serie': 'B001', 'numero': 2,
            'cliente': ('1', '12345678', 'JUAN PEREZ LOPEZ', 'JR. LIMA 789'),
            'items': [
                ('P004', 'Leche Gloria 1L', 4, 'NIU', 4.20),
                ('P005', 'Pan Integral 500g', 6, 'NIU', 2.80),
                ('P006', 'Fideos Don Vittorio 1kg', 2, 'NIU', 6.50)
            ]
        }
    ]
    
    fecha_base = datetime.now()
    
    for comp in comprobantes:
        tipo = comp['tipo']
        serie = comp['serie']
        numero = comp['numero']
        cli_tipo, cli_doc, cli_nom, cli_dir = comp['cliente']
        
        # Calcular totales del comprobante
        subtotal = 0
        igv_total = 0
        
        for item_codigo, item_desc, cant, unidad, precio_unit in comp['items']:
            # Precios
            precio_con_igv = precio_unit
            valor_sin_igv = round(precio_con_igv / 1.18, 2)
            subtotal_item = round(valor_sin_igv * cant, 2)
            igv_item = round(subtotal_item * 0.18, 2)
            total_item = subtotal_item + igv_item
            
            subtotal += subtotal_item
            igv_total += igv_item
            
            # Fila
            ws.append([
                tipo, serie, numero, fecha_base.strftime('%Y-%m-%d'), 'PEN',
                cli_tipo, cli_doc, cli_nom, cli_dir,
                item_codigo, item_desc, cant, unidad,
                precio_con_igv, valor_sin_igv,
                subtotal_item, igv_item, total_item,
                '10', '10000000',
                '', '', '', '',  # Se llenan solo en última fila
                '', '',
                '', '',
                ''  # Estado vacío = PENDIENTE
            ])
        
        # Última fila del comprobante con totales
        ultima_fila = list(ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row))[0]
        ultima_fila[20].value = round(subtotal, 2)  # venta_subtotal
        ultima_fila[21].value = 0.00  # venta_exonerada
        ultima_fila[22].value = 0.00  # venta_inafecta
        ultima_fila[23].value = 0.00  # venta_icbper
        ultima_fila[24].value = round(igv_total, 2)  # venta_igv
        ultima_fila[25].value = round(subtotal + igv_total, 2)  # venta_total
        ultima_fila[26].value = 'Contado'  # pago_forma
        ultima_fila[27].value = round(subtotal + igv_total, 2)  # pago_monto
        ultima_fila[28].value = 'PENDIENTE'  # estado
        
        fecha_base += timedelta(hours=2)
    
    # Guardar
    filename = 'ventas_pos_prueba.xlsx'
    wb.save(filename)
    print(f"✅ Archivo generado: {filename}")
    print(f"📊 Comprobantes: {len(comprobantes)}")
    print(f"   - 2 Boletas (B001-0001, B001-0002)")
    print(f"   - 1 Factura (F001-0001)")

if __name__ == '__main__':
    generar_excel_prueba()
