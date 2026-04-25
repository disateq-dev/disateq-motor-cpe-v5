"""
source_explorer.py
==================
Herramienta de descubrimiento de fuentes de datos — DisateQ CPE™ v4.0

Analiza una fuente de datos (DBF, SQL, XLSX, CSV) y genera:
1. Reporte de estructura (tablas, campos, tipos, muestras)
2. Borrador de contrato YAML para el adaptador genérico

Uso:
    from src.tools.source_explorer import SourceExplorer
    explorer = SourceExplorer()
    reporte  = explorer.explorar(tipo='dbf', ruta='D:/SISTEMA/datos/')
    explorer.guardar_contrato(reporte, 'config/contratos/cliente_xyz.yaml')
"""

import os
import yaml
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional


class SourceExplorer:
    """
    Explora fuentes de datos y genera contratos YAML para el Motor CPE.
    Soporta: DBF, SQL Server, MySQL, PostgreSQL, XLSX, CSV.
    """

    TIPOS_SOPORTADOS = ['dbf', 'sqlserver', 'mysql', 'postgres', 'xlsx', 'csv']

    def __init__(self):
        self.reporte = {}

    # ================================================================
    # ENTRY POINT
    # ================================================================

    def explorar(self,
                 tipo: str,
                 ruta: str = None,
                 servidor: str = None,
                 base_datos: str = None,
                 usuario: str = None,
                 clave: str = None,
                 puerto: int = None) -> Dict:
        """
        Explora una fuente de datos y retorna reporte completo.

        Args:
            tipo:       'dbf' | 'sqlserver' | 'mysql' | 'postgres' | 'xlsx' | 'csv'
            ruta:       Ruta a carpeta (DBF) o archivo (XLSX/CSV)
            servidor:   Servidor SQL
            base_datos: Base de datos SQL
            usuario:    Usuario SQL
            clave:      Contraseña SQL
            puerto:     Puerto SQL

        Returns:
            Dict con reporte completo de la fuente
        """
        tipo = tipo.lower().strip()
        if tipo not in self.TIPOS_SOPORTADOS:
            raise ValueError(f"Tipo no soportado: {tipo}. Usa: {self.TIPOS_SOPORTADOS}")

        print(f"\n{'='*60}")
        print(f"  DisateQ CPE™ — Source Explorer")
        print(f"  Analizando fuente: {tipo.upper()}")
        print(f"{'='*60}\n")

        if tipo == 'dbf':
            reporte = self._explorar_dbf(ruta)
        elif tipo in ('sqlserver', 'mysql', 'postgres'):
            reporte = self._explorar_sql(tipo, servidor, base_datos, usuario, clave, puerto)
        elif tipo == 'xlsx':
            reporte = self._explorar_xlsx(ruta)
        elif tipo == 'csv':
            reporte = self._explorar_csv(ruta)
        else:
            raise ValueError(f"Tipo no implementado: {tipo}")

        reporte['meta'] = {
            'tipo':          tipo,
            'fecha_analisis': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'version_explorer': '4.0'
        }

        self.reporte = reporte
        self._imprimir_reporte(reporte)
        return reporte

    # ================================================================
    # DBF
    # ================================================================

    def _explorar_dbf(self, ruta: str) -> Dict:
        """Explora carpeta DBF y analiza todos los archivos .dbf."""
        from dbfread import DBF

        path = Path(ruta)
        if not path.exists():
            raise FileNotFoundError(f"Ruta no encontrada: {ruta}")

        archivos = list(path.glob('*.dbf'))
        if not archivos:
            raise FileNotFoundError(f"No se encontraron archivos .dbf en: {ruta}")

        print(f"📁 Carpeta: {ruta}")
        print(f"   Archivos DBF encontrados: {len(archivos)}\n")

        tablas = {}
        for dbf_path in sorted(archivos):
            nombre = dbf_path.stem.lower()
            try:
                dbf = DBF(str(dbf_path), encoding='latin1',
                          ignore_missing_memofile=True, raw=True)

                campos = []
                for field in dbf.fields:
                    campos.append({
                        'nombre': field.name,
                        'tipo':   field.type,
                        'longitud': field.length,
                        'decimales': getattr(field, 'decimal_count', 0)
                    })

                # Contar registros y tomar muestra
                total = 0
                muestra = []
                for i, record in enumerate(dbf):
                    total += 1
                    if i < 3:
                        muestra.append({
                            k: v.decode('latin1').strip() if isinstance(v, bytes) else str(v)
                            for k, v in record.items()
                        })

                tablas[nombre] = {
                    'archivo':  dbf_path.name,
                    'campos':   campos,
                    'total_registros': total,
                    'muestra':  muestra
                }
                print(f"   ✅ {dbf_path.name:<35} {len(campos):>3} campos  {total:>8} registros")

            except Exception as e:
                print(f"   ⚠️  {dbf_path.name:<35} Error: {e}")

        return {
            'tipo':   'dbf',
            'ruta':   str(ruta),
            'tablas': tablas
        }

    # ================================================================
    # SQL
    # ================================================================

    def _explorar_sql(self, tipo: str, servidor: str, base_datos: str,
                      usuario: str, clave: str, puerto: int) -> Dict:
        """Explora base de datos SQL."""
        conn = self._conectar_sql(tipo, servidor, base_datos, usuario, clave, puerto)

        print(f"🗄️  Servidor: {servidor}")
        print(f"   Base de datos: {base_datos}\n")

        tablas = self._listar_tablas_sql(conn, tipo, base_datos)
        resultado = {}

        for tabla in tablas:
            try:
                campos  = self._describir_tabla_sql(conn, tabla, tipo)
                total   = self._contar_registros_sql(conn, tabla)
                muestra = self._muestra_sql(conn, tabla)

                resultado[tabla.lower()] = {
                    'tabla':   tabla,
                    'campos':  campos,
                    'total_registros': total,
                    'muestra': muestra
                }
                print(f"   ✅ {tabla:<40} {len(campos):>3} campos  {total:>8} registros")
            except Exception as e:
                print(f"   ⚠️  {tabla:<40} Error: {e}")

        conn.close()
        return {
            'tipo':       tipo,
            'servidor':   servidor,
            'base_datos': base_datos,
            'tablas':     resultado
        }

    def _conectar_sql(self, tipo, servidor, base_datos, usuario, clave, puerto):
        if tipo == 'sqlserver':
            import pyodbc
            puerto_str = f",{puerto}" if puerto else ""
            conn_str = (
                f"DRIVER={{ODBC Driver 17 for SQL Server}};"
                f"SERVER={servidor}{puerto_str};"
                f"DATABASE={base_datos};"
                f"UID={usuario};PWD={clave}"
            )
            return pyodbc.connect(conn_str)

        elif tipo == 'mysql':
            import mysql.connector
            return mysql.connector.connect(
                host=servidor, database=base_datos,
                user=usuario, password=clave,
                port=puerto or 3306
            )

        elif tipo == 'postgres':
            import psycopg2
            return psycopg2.connect(
                host=servidor, dbname=base_datos,
                user=usuario, password=clave,
                port=puerto or 5432
            )

    def _listar_tablas_sql(self, conn, tipo, base_datos) -> List[str]:
        cursor = conn.cursor()
        if tipo == 'sqlserver':
            cursor.execute("SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_TYPE='BASE TABLE' ORDER BY TABLE_NAME")
        elif tipo == 'mysql':
            cursor.execute(f"SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA='{base_datos}' ORDER BY TABLE_NAME")
        elif tipo == 'postgres':
            cursor.execute("SELECT tablename FROM pg_tables WHERE schemaname='public' ORDER BY tablename")
        return [row[0] for row in cursor.fetchall()]

    def _describir_tabla_sql(self, conn, tabla, tipo) -> List[Dict]:
        cursor = conn.cursor()
        if tipo == 'sqlserver':
            cursor.execute(f"""
                SELECT COLUMN_NAME, DATA_TYPE, CHARACTER_MAXIMUM_LENGTH, NUMERIC_PRECISION
                FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME='{tabla}' ORDER BY ORDINAL_POSITION
            """)
        elif tipo == 'mysql':
            cursor.execute(f"DESCRIBE `{tabla}`")
        elif tipo == 'postgres':
            cursor.execute(f"""
                SELECT column_name, data_type, character_maximum_length, numeric_precision
                FROM information_schema.columns WHERE table_name='{tabla}' ORDER BY ordinal_position
            """)
        return [{'nombre': row[0], 'tipo': str(row[1]), 'longitud': row[2]} for row in cursor.fetchall()]

    def _contar_registros_sql(self, conn, tabla) -> int:
        try:
            cursor = conn.cursor()
            cursor.execute(f"SELECT COUNT(*) FROM [{tabla}]")
            return cursor.fetchone()[0]
        except:
            return 0

    def _muestra_sql(self, conn, tabla) -> List[Dict]:
        try:
            cursor = conn.cursor()
            cursor.execute(f"SELECT TOP 3 * FROM [{tabla}]")
            cols = [d[0] for d in cursor.description]
            return [dict(zip(cols, [str(v) for v in row])) for row in cursor.fetchall()]
        except:
            return []

    # ================================================================
    # XLSX
    # ================================================================

    def _explorar_xlsx(self, ruta: str) -> Dict:
        """Explora archivo Excel."""
        import openpyxl
        path = Path(ruta)
        if not path.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {ruta}")

        print(f"📊 Archivo: {ruta}\n")

        wb     = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        hojas  = {}

        for nombre_hoja in wb.sheetnames:
            try:
                ws     = wb[nombre_hoja]
                filas  = list(ws.iter_rows(values_only=True))
                if not filas:
                    continue

                encabezados = [str(c) if c is not None else f'COL_{i}' for i, c in enumerate(filas[0])]
                total = len(filas) - 1

                muestra = []
                for fila in filas[1:4]:
                    muestra.append(dict(zip(encabezados, [str(v) if v is not None else '' for v in fila])))

                campos = []
                for i, enc in enumerate(encabezados):
                    tipo_detectado = 'texto'
                    for fila in filas[1:10]:
                        val = fila[i] if i < len(fila) else None
                        if isinstance(val, (int, float)):
                            tipo_detectado = 'numerico'
                            break
                    campos.append({'nombre': enc, 'tipo': tipo_detectado, 'longitud': None})

                hojas[nombre_hoja] = {
                    'hoja':    nombre_hoja,
                    'campos':  campos,
                    'total_registros': total,
                    'muestra': muestra
                }
                print(f"   ✅ Hoja: {nombre_hoja:<30} {len(campos):>3} columnas  {total:>8} filas")

            except Exception as e:
                print(f"   ⚠️  Hoja: {nombre_hoja:<30} Error: {e}")

        wb.close()
        return {'tipo': 'xlsx', 'ruta': str(ruta), 'tablas': hojas}

    # ================================================================
    # CSV
    # ================================================================

    def _explorar_csv(self, ruta: str) -> Dict:
        """Explora archivo CSV."""
        import csv
        path = Path(ruta)
        if not path.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {ruta}")

        print(f"📄 Archivo: {ruta}\n")

        with open(ruta, encoding='utf-8', errors='replace') as f:
            reader  = csv.DictReader(f)
            campos  = [{'nombre': c, 'tipo': 'texto', 'longitud': None} for c in (reader.fieldnames or [])]
            muestra = []
            total   = 0
            for row in reader:
                total += 1
                if total <= 3:
                    muestra.append(dict(row))

        nombre = path.stem
        print(f"   ✅ {path.name:<35} {len(campos):>3} campos  {total:>8} registros")

        return {
            'tipo':   'csv',
            'ruta':   str(ruta),
            'tablas': {nombre: {'archivo': path.name, 'campos': campos,
                                'total_registros': total, 'muestra': muestra}}
        }

    # ================================================================
    # IMPRIMIR REPORTE
    # ================================================================

    def _imprimir_reporte(self, reporte: Dict):
        tablas = reporte.get('tablas', {})
        print(f"\n{'='*60}")
        print(f"  RESUMEN — {len(tablas)} tablas/archivos analizados")
        print(f"{'='*60}")
        total_campos = sum(len(t.get('campos', [])) for t in tablas.values())
        total_registros = sum(t.get('total_registros', 0) for t in tablas.values())
        print(f"  Total campos:    {total_campos}")
        print(f"  Total registros: {total_registros:,}")
        print(f"{'='*60}\n")

    # ================================================================
    # GENERAR CONTRATO YAML
    # ================================================================

    def generar_contrato(self,
                         reporte: Dict = None,
                         tabla_comprobantes: str = None,
                         tabla_items: str = None,
                         tabla_anulaciones: str = None) -> Dict:
        """
        Genera borrador de contrato YAML a partir del reporte.

        Args:
            reporte:              Reporte generado por explorar()
            tabla_comprobantes:   Nombre de la tabla principal de comprobantes
            tabla_items:          Nombre de la tabla de items/detalle
            tabla_anulaciones:    Nombre de la tabla de anulaciones (opcional)

        Returns:
            Dict con estructura del contrato YAML
        """
        if reporte is None:
            reporte = self.reporte
        if not reporte:
            raise ValueError("No hay reporte. Ejecuta explorar() primero.")

        tipo   = reporte.get('tipo', 'dbf')
        tablas = reporte.get('tablas', {})

        # Auto-detectar tablas si no se especifican
        if not tabla_comprobantes:
            tabla_comprobantes = self._detectar_tabla_comprobantes(tablas)
        if not tabla_items:
            tabla_items = self._detectar_tabla_items(tablas)

        contrato = {
            'version': '1.0',
            'generado': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'motor': 'DisateQ CPE v4.0',

            'fuente': self._generar_bloque_fuente(tipo, reporte),

            'comprobantes': self._generar_bloque_mapeo(
                tablas.get(tabla_comprobantes, {}),
                tabla_comprobantes,
                tipo='comprobantes'
            ),

            'items': self._generar_bloque_mapeo(
                tablas.get(tabla_items, {}),
                tabla_items,
                tipo='items'
            ) if tabla_items else None,

            'anulaciones': self._generar_bloque_mapeo(
                tablas.get(tabla_anulaciones, {}),
                tabla_anulaciones,
                tipo='anulaciones'
            ) if tabla_anulaciones else None,

            'transformaciones': {
                'tipo_doc': {
                    '__comentario': 'Mapear valor del sistema origen al codigo SUNAT',
                    'F': '01',
                    'B': '03',
                    'T': '03'
                },
                'fecha': {
                    'formato_origen': 'YYYYMMDD',
                    '__opciones': ['YYYYMMDD', 'DD/MM/YYYY', 'YYYY-MM-DD', 'DD-MM-YYYY']
                },
                'moneda': {
                    '__comentario': 'Mapear moneda del sistema al codigo SUNAT',
                    'S': 'PEN',
                    'D': 'USD'
                }
            }
        }

        # Limpiar nulos
        contrato = {k: v for k, v in contrato.items() if v is not None}
        return contrato

    def _generar_bloque_fuente(self, tipo: str, reporte: Dict) -> Dict:
        if tipo == 'dbf':
            return {
                'tipo':   'dbf',
                'ruta':   reporte.get('ruta', ''),
                '__nota': 'Ruta a la carpeta que contiene los archivos .dbf'
            }
        elif tipo in ('sqlserver', 'mysql', 'postgres'):
            return {
                'tipo':       tipo,
                'servidor':   reporte.get('servidor', ''),
                'base_datos': reporte.get('base_datos', ''),
                'usuario':    '',
                'clave':      '',
                'puerto':     {'sqlserver': 1433, 'mysql': 3306, 'postgres': 5432}.get(tipo)
            }
        elif tipo == 'xlsx':
            return {
                'tipo':   'xlsx',
                'ruta':   reporte.get('ruta', ''),
                '__nota': 'Ruta al archivo .xlsx'
            }
        elif tipo == 'csv':
            return {
                'tipo':      'csv',
                'ruta':      reporte.get('ruta', ''),
                'separador': ',',
                'encoding':  'utf-8'
            }
        return {'tipo': tipo}

    def _generar_bloque_mapeo(self, tabla_info: Dict, nombre_tabla: str, tipo: str) -> Dict:
        """Genera bloque de mapeo con campos detectados y sugerencias."""
        if not tabla_info:
            return {
                'tabla': nombre_tabla or 'COMPLETAR',
                'filtro': '__COMPLETAR: condicion para registros pendientes',
                'campos': self._campos_requeridos(tipo)
            }

        campos_origen = [c['nombre'] for c in tabla_info.get('campos', [])]

        # Sugerir mapeos por similitud de nombres
        mapeo = self._sugerir_mapeo(campos_origen, tipo)

        bloque = {
            'tabla':  nombre_tabla,
            'filtro': '__COMPLETAR',
            'campos': mapeo
        }

        if tipo == 'items':
            bloque['join'] = '__COMPLETAR: campo de union con tabla comprobantes'

        return bloque

    def _campos_requeridos(self, tipo: str) -> Dict:
        if tipo == 'comprobantes':
            return {
                'tipo_doc':      '__CAMPO_ORIGEN',
                'serie':         '__CAMPO_ORIGEN',
                'numero':        '__CAMPO_ORIGEN',
                'fecha':         '__CAMPO_ORIGEN',
                'total':         '__CAMPO_ORIGEN',
                'ruc_cliente':   '__CAMPO_ORIGEN (opcional boletas)',
                'nombre_cliente':'__CAMPO_ORIGEN',
                'direccion':     '__CAMPO_ORIGEN (opcional)'
            }
        elif tipo == 'items':
            return {
                'codigo':      '__CAMPO_ORIGEN',
                'descripcion': '__CAMPO_ORIGEN',
                'cantidad':    '__CAMPO_ORIGEN',
                'precio':      '__CAMPO_ORIGEN',
                'subtotal':    '__CAMPO_ORIGEN',
                'igv':         '__CAMPO_ORIGEN',
                'total':       '__CAMPO_ORIGEN'
            }
        elif tipo == 'anulaciones':
            return {
                'tipo_doc':        '__CAMPO_ORIGEN',
                'serie':           '__CAMPO_ORIGEN',
                'numero':          '__CAMPO_ORIGEN',
                'fecha_emision':   '__CAMPO_ORIGEN',
                'motivo':          '__CAMPO_ORIGEN'
            }
        return {}

    def _sugerir_mapeo(self, campos_origen: List[str], tipo: str) -> Dict:
        """Sugiere mapeo por similitud de nombres de campo."""
        campos_lower = {c.lower(): c for c in campos_origen}

        # Patrones de búsqueda por campo estándar
        patrones = {
            'comprobantes': {
                'tipo_doc':       ['tipo_factu', 'tipo_doc', 'tipo_comp', 'tipo', 'cod_tipo'],
                'serie':          ['serie_fact', 'serie', 'num_serie', 'cod_serie'],
                'numero':         ['numero_fac', 'numero', 'num_doc', 'nro_doc', 'nrodoc'],
                'fecha':          ['fecha_fact', 'fecha_doc', 'fecha_emi', 'fecha', 'fec_emision'],
                'total':          ['total_fact', 'total', 'monto_total', 'importe_total'],
                'ruc_cliente':    ['ruc_client', 'ruc', 'num_doc_cli', 'documento'],
                'nombre_cliente': ['nombre_cli', 'razon_soci', 'cliente', 'nom_cliente'],
                'direccion':      ['direccion', 'dir_client', 'domicilio']
            },
            'items': {
                'codigo':      ['codigo_pro', 'cod_prod', 'codigo', 'cod_articulo'],
                'descripcion': ['descripcio', 'descripcion', 'nombre_pro', 'detalle'],
                'cantidad':    ['cantidad_p', 'cantidad', 'cant', 'fraccion_p'],
                'precio':      ['precio_uni', 'precio', 'precio_venta', 'precio_con'],
                'subtotal':    ['monto_pedi', 'subtotal', 'importe', 'valor'],
                'igv':         ['igv_pedido', 'igv', 'impuesto'],
                'total':       ['real_pedid', 'total', 'total_item', 'precio_total']
            },
            'anulaciones': {
                'tipo_doc':      ['tipo_factu', 'tipo_doc', 'tipo'],
                'serie':         ['serie_nota', 'serie_fact', 'serie'],
                'numero':        ['numero_not', 'numero_fac', 'numero'],
                'fecha_emision': ['fecha_nota', 'fecha_fact', 'fecha'],
                'motivo':        ['tipo_motiv', 'motivo', 'descripcion']
            }
        }

        mapeo_sugerido = {}
        patrones_tipo  = patrones.get(tipo, {})

        for campo_std, candidatos in patrones_tipo.items():
            encontrado = None
            for candidato in candidatos:
                if candidato in campos_lower:
                    encontrado = campos_lower[candidato]
                    break
            mapeo_sugerido[campo_std] = encontrado if encontrado else f'__COMPLETAR (opciones: {", ".join(campos_origen[:5])})'

        return mapeo_sugerido

    def _detectar_tabla_comprobantes(self, tablas: Dict) -> Optional[str]:
        """Intenta detectar la tabla principal de comprobantes."""
        candidatos = ['enviosffee', 'facturas', 'ventas', 'comprobantes',
                      'factura', 'venta', 'documento', 'cabecera']
        for c in candidatos:
            if c in tablas:
                return c
        # Mayor cantidad de registros
        if tablas:
            return max(tablas.keys(), key=lambda t: tablas[t].get('total_registros', 0))
        return None

    def _detectar_tabla_items(self, tablas: Dict) -> Optional[str]:
        """Intenta detectar la tabla de items/detalle."""
        candidatos = ['detalleventa', 'detalle_ventas', 'items', 'detalle',
                      'lineas', 'detalles', 'factura_det', 'venta_det']
        for c in candidatos:
            if c in tablas:
                return c
        return None

    # ================================================================
    # GUARDAR CONTRATO
    # ================================================================

    def guardar_contrato(self, contrato: Dict, ruta_destino: str) -> str:
        """Guarda el contrato YAML en disco."""
        path = Path(ruta_destino)
        path.parent.mkdir(parents=True, exist_ok=True)

        # Agregar comentario de encabezado
        header = (
            "# ============================================================\n"
            "# CONTRATO DE DATOS DisateQ CPE™ v4.0\n"
            "# Generado automáticamente por Source Explorer\n"
            f"# Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            "#\n"
            "# INSTRUCCIONES:\n"
            "# 1. Reemplaza los valores __COMPLETAR con los campos reales\n"
            "# 2. Ajusta los filtros según el estado 'pendiente' en tu sistema\n"
            "# 3. Verifica el mapeo de tipo_doc con los valores de tu sistema\n"
            "# 4. Guarda y prueba con: python -m src.tools.source_explorer --validar\n"
            "# ============================================================\n\n"
        )

        with open(path, 'w', encoding='utf-8') as f:
            f.write(header)
            yaml.dump(contrato, f, allow_unicode=True,
                      default_flow_style=False, sort_keys=False)

        print(f"\n✅ Contrato guardado: {path}")
        print(f"   Edita los campos marcados con __COMPLETAR\n")
        return str(path)

    def guardar_reporte(self, reporte: Dict, ruta_destino: str) -> str:
        """Guarda el reporte de exploración en YAML para revisión del técnico."""
        path = Path(ruta_destino)
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, 'w', encoding='utf-8') as f:
            yaml.dump(reporte, f, allow_unicode=True,
                      default_flow_style=False, sort_keys=False)
        print(f"✅ Reporte guardado: {path}")
        return str(path)


# ================================================================
# CLI
# ================================================================

def main():
    import argparse

    parser = argparse.ArgumentParser(
        description='DisateQ CPE™ Source Explorer — Analiza fuentes de datos'
    )
    parser.add_argument('--tipo',   required=True,
                        choices=SourceExplorer.TIPOS_SOPORTADOS,
                        help='Tipo de fuente')
    parser.add_argument('--ruta',   help='Ruta a carpeta DBF o archivo XLSX/CSV')
    parser.add_argument('--servidor', help='Servidor SQL')
    parser.add_argument('--bd',     help='Base de datos SQL')
    parser.add_argument('--usuario', help='Usuario SQL')
    parser.add_argument('--clave',  help='Contraseña SQL')
    parser.add_argument('--puerto', type=int, help='Puerto SQL')
    parser.add_argument('--salida', default='output/explorer',
                        help='Directorio de salida (default: output/explorer)')
    parser.add_argument('--tabla-comp',  help='Tabla de comprobantes')
    parser.add_argument('--tabla-items', help='Tabla de items')
    parser.add_argument('--tabla-anul',  help='Tabla de anulaciones')

    args = parser.parse_args()

    explorer = SourceExplorer()
    reporte  = explorer.explorar(
        tipo      = args.tipo,
        ruta      = args.ruta,
        servidor  = args.servidor,
        base_datos= args.bd,
        usuario   = args.usuario,
        clave     = args.clave,
        puerto    = args.puerto
    )

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    explorer.guardar_reporte(reporte, f"{args.salida}/reporte_{ts}.yaml")

    contrato = explorer.generar_contrato(
        reporte,
        tabla_comprobantes = args.tabla_comp,
        tabla_items        = args.tabla_items,
        tabla_anulaciones  = args.tabla_anul
    )
    explorer.guardar_contrato(contrato, f"{args.salida}/contrato_{ts}.yaml")


if __name__ == '__main__':
    main()
