# ══════════════════════════════════════════════════════════════════
#  DisateQ Motor CPE v5.0  —  wizard_service.py
#  TASK-006: _build_contrato_yaml estructura flag_lectura/flag_escritura
#            correcta para GenericAdapter
# ══════════════════════════════════════════════════════════════════

from __future__ import annotations
import traceback
from pathlib import Path
import yaml


def _ruta_config() -> Path:
    here = Path(__file__).resolve()
    for p in [here.parent.parent.parent, here.parent.parent]:
        if (p / "config").is_dir():
            return p / "config"
    return here.parent.parent.parent / "config"


# ══════════════════════════════════════════════════════════════════
#  TEST FUENTE — paso 3
# ══════════════════════════════════════════════════════════════════

def test_fuente(fuente: dict) -> dict:
    tipo = fuente.get("tipo", "").lower()
    try:
        if tipo == "dbf":     return _test_dbf(fuente)
        elif tipo == "excel": return _test_excel(fuente)
        elif tipo == "csv":   return _test_csv(fuente)
        elif tipo in ("sqlserver", "mysql", "postgresql"):
            return _test_db(fuente)
        elif tipo == "access": return _test_access(fuente)
        else:
            return {"ok": False, "error": f"Tipo no soportado: '{tipo}'"}
    except Exception as exc:
        return {"ok": False, "error": str(exc)}


def _test_dbf(fuente: dict) -> dict:
    try:
        from dbfread import DBF as DbfReader
    except ImportError:
        return {"ok": False, "error": "dbfread no instalado. pip install dbfread"}

    ruta = fuente.get("ruta", "").strip()
    if not ruta:
        return {"ok": False, "error": "Ruta de carpeta no especificada"}
    carpeta = Path(ruta)
    if not carpeta.exists():
        return {"ok": False, "error": f"La ruta no existe: {ruta}"}

    dbfs = sorted({p.stem.lower(): p for p in
                   list(carpeta.glob("*.dbf")) +
                   list(carpeta.glob("*.DBF"))}.values())
    if not dbfs:
        return {"ok": False, "error": f"No se encontraron archivos .DBF en: {ruta}"}

    archivos = []
    for dbf_path in dbfs:
        try:
            t = DbfReader(str(dbf_path), encoding="latin-1", load=False)
            archivos.append({
                "archivo":      dbf_path.name,
                "campos":       len(t.fields),
                "lista_campos": ", ".join(f.name for f in t.fields[:6]) +
                                ("..." if len(t.fields) > 6 else ""),
            })
        except Exception:
            archivos.append({"archivo": dbf_path.name, "campos": 0,
                             "lista_campos": "(no se pudo leer)"})

    return {
        "ok":      True,
        "mensaje": f"{len(dbfs)} archivo(s) DBF encontrado(s) en la carpeta.",
        "columnas": ["archivo", "campos", "lista_campos"],
        "filas":    archivos,
        "total_registros": len(archivos),
    }


def _test_excel(fuente: dict) -> dict:
    try:
        import openpyxl
    except ImportError:
        return {"ok": False, "error": "openpyxl no instalado."}
    ruta = Path(fuente.get("ruta", "").strip())
    if not ruta.exists():
        return {"ok": False, "error": f"Archivo no encontrado: {ruta}"}
    wb   = openpyxl.load_workbook(str(ruta), read_only=True, data_only=True)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True, max_row=6))
    wb.close()
    if not rows:
        return {"ok": False, "error": "Archivo Excel vacío"}
    cols  = [str(c) if c is not None else f"Col{i}" for i, c in enumerate(rows[0])]
    filas = [{cols[j]: str(v) for j, v in enumerate(r)} for r in rows[1:6]]
    return {"ok": True, "mensaje": f"Excel OK — {ws.max_row - 1} filas.",
            "columnas": cols[:12], "filas": filas, "total_registros": ws.max_row - 1}


def _test_csv(fuente: dict) -> dict:
    import csv as _csv
    ruta = Path(fuente.get("ruta", "").strip())
    if not ruta.exists():
        return {"ok": False, "error": f"Archivo no encontrado: {ruta}"}
    with open(ruta, newline="", encoding="latin-1", errors="replace") as f:
        reader = _csv.DictReader(f)
        cols   = reader.fieldnames or []
        filas  = [row for i, row in enumerate(reader) if i < 5]
    return {"ok": True, "mensaje": f"CSV OK — {len(cols)} columnas.",
            "columnas": list(cols)[:12], "filas": filas, "total_registros": len(filas)}


def _test_db(fuente: dict) -> dict:
    tipo = fuente.get("tipo"); host = fuente.get("host", "localhost")
    puerto = fuente.get("puerto", ""); database = fuente.get("database", "")
    usuario = fuente.get("usuario", ""); password = fuente.get("password", "")
    try:
        if tipo == "sqlserver":
            import pyodbc
            cs   = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={host},{puerto or 1433};DATABASE={database};UID={usuario};PWD={password}"
            conn = pyodbc.connect(cs, timeout=5); cur = conn.cursor()
            cur.execute("SELECT TOP 10 TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_TYPE='BASE TABLE'")
        elif tipo == "mysql":
            import mysql.connector
            conn = mysql.connector.connect(host=host, port=int(puerto or 3306),
                database=database, user=usuario, password=password, connection_timeout=5)
            cur = conn.cursor(); cur.execute("SHOW TABLES")
        elif tipo == "postgresql":
            import psycopg2
            conn = psycopg2.connect(host=host, port=int(puerto or 5432),
                dbname=database, user=usuario, password=password, connect_timeout=5)
            cur = conn.cursor()
            cur.execute("SELECT table_name FROM information_schema.tables WHERE table_schema='public' LIMIT 10")
        else:
            return {"ok": False, "error": f"DB tipo no soportado: {tipo}"}
        tablas = [row[0] for row in cur.fetchall()]
        cur.close(); conn.close()
        return {"ok": True, "mensaje": f"Conexión OK — {len(tablas)} tabla(s).",
                "columnas": ["tabla"], "filas": [{"tabla": t} for t in tablas],
                "total_registros": len(tablas)}
    except Exception as exc:
        return {"ok": False, "error": f"No se pudo conectar a {tipo}: {exc}"}


def _test_access(fuente: dict) -> dict:
    try:
        import pyodbc
    except ImportError:
        return {"ok": False, "error": "pyodbc no instalado."}
    ruta = fuente.get("ruta", "").strip()
    if not Path(ruta).exists():
        return {"ok": False, "error": f"Archivo no encontrado: {ruta}"}
    try:
        conn   = pyodbc.connect(f"DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={ruta};")
        cur    = conn.cursor()
        tablas = [r.table_name for r in cur.tables(tableType="TABLE")]
        cur.close(); conn.close()
        return {"ok": True, "mensaje": f"Access OK — {len(tablas)} tabla(s).",
                "columnas": ["tabla"], "filas": [{"tabla": t} for t in tablas],
                "total_registros": len(tablas)}
    except Exception as exc:
        return {"ok": False, "error": f"Error Access: {exc}"}


# ══════════════════════════════════════════════════════════════════
#  ANALIZAR FUENTE — heurística DBF → contrato
# ══════════════════════════════════════════════════════════════════

def analizar_fuente(fuente: dict) -> dict:
    tipo = fuente.get("tipo", "").lower()
    if tipo == "dbf":
        try:
            from src.tools.wizard_mapper import mapear_dbf
            result = mapear_dbf(fuente.get("ruta", "").strip())
            result["metodo"] = "heuristica"
            return result
        except Exception as exc:
            return {"ok": False, "error": str(exc), "metodo": "manual",
                    "score_global": 0, "contrato": {}, "scores": {}, "sin_resolver": []}
    return {"ok": True, "score_global": 0, "metodo": "manual",
            "contrato": {}, "scores": {}, "sin_resolver": [], "tablas_analizadas": 0}


# ══════════════════════════════════════════════════════════════════
#  PROBAR MAPEO — lee 5 registros reales con el contrato actual
# ══════════════════════════════════════════════════════════════════

def probar_mapeo(fuente: dict, contrato: dict) -> dict:
    tipo = fuente.get("tipo", "").lower()
    if tipo != "dbf":
        return {"ok": False, "error": "Prueba disponible solo para DBF por ahora."}
    try:
        from dbfread import DBF as DbfReader
    except ImportError:
        return {"ok": False, "error": "dbfread no instalado."}

    ruta   = fuente.get("ruta", "").strip()
    tabla  = contrato.get("tabla", "").strip()
    campos = contrato.get("campos", {})
    flag_c = contrato.get("flag_campo", "").strip()

    if not tabla:
        return {"ok": False, "error": "Tabla de comprobantes no definida."}

    dbf_path = Path(ruta) / f"{tabla}.dbf"
    if not dbf_path.exists():
        dbf_path = Path(ruta) / f"{tabla.upper()}.DBF"
    if not dbf_path.exists():
        return {"ok": False, "error": f"Archivo no encontrado: {tabla}.dbf en {ruta}"}

    try:
        t = DbfReader(str(dbf_path), encoding="latin-1", load=False)
        cols_cpe = {
            "serie":          campos.get("serie", ""),
            "numero":         campos.get("numero", ""),
            "tipo_doc":       campos.get("tipo_doc", ""),
            "fecha":          campos.get("fecha", ""),
            "ruc_cliente":    campos.get("ruc_cliente", ""),
            "nombre_cliente": campos.get("nombre_cliente", ""),
            "total":          campos.get("total", ""),
        }
        if flag_c:
            cols_cpe["flag"] = flag_c
        cols_mostrar = [k for k, v in cols_cpe.items() if v]
        filas = []
        for i, rec in enumerate(t):
            if i >= 5: break
            fila = {}
            for col_cpe in cols_mostrar:
                campo_real = cols_cpe[col_cpe]
                v = rec.get(campo_real, "")
                fila[col_cpe] = str(v).strip() if v is not None else ""
            filas.append(fila)
        if not filas:
            return {"ok": False, "error": "La tabla está vacía."}
        return {"ok": True, "columnas": cols_mostrar, "filas": filas, "tabla": tabla}
    except Exception as exc:
        return {"ok": False, "error": f"Error leyendo {tabla}.dbf: {exc}"}


# ══════════════════════════════════════════════════════════════════
#  GUARDAR WIZARD → YAML compatible con ClientLoader
# ══════════════════════════════════════════════════════════════════

def guardar_wizard(payload: dict) -> dict:
    """
    Genera config/clientes/{id}.yaml y config/contratos/{id}.yaml
    en el formato que ClientLoader y GenericAdapter esperan.
    """
    try:
        cfg_root   = _ruta_config()
        cliente    = payload["cliente"]
        fuente     = payload["fuente"]
        contrato   = payload["contrato"]
        series_raw = payload["series"]       # { '01': [{serie,correlativo_inicio}], ... }
        creds      = payload["credenciales"] # { nombre, tipo, usuario, token, endpoints: [{tipo,url}] }
        cliente_id = cliente["cliente_id"]

        # ── YAML cliente (formato ClientLoader) ─────────────────
        cliente_yaml = _build_cliente_yaml(cliente, fuente, series_raw, creds)
        d = cfg_root / "clientes"
        d.mkdir(parents=True, exist_ok=True)
        _write_yaml(d / f"{cliente_id}.yaml", cliente_yaml)

        # ── YAML contrato (formato GenericAdapter) ───────────────
        contrato_yaml = _build_contrato_yaml(cliente_id, fuente, contrato)
        d = cfg_root / "contratos"
        d.mkdir(parents=True, exist_ok=True)
        _write_yaml(d / f"{cliente_id}.yaml", contrato_yaml)

        return {"ok": True, "cliente_id": cliente_id}
    except Exception as exc:
        traceback.print_exc()
        return {"ok": False, "error": str(exc)}


def _build_cliente_yaml(cliente: dict, fuente: dict, series_raw: dict, creds: dict) -> dict:
    """
    Genera YAML en el mismo formato que farmacia_central.yaml
    para que ClientLoader lo lea sin cambios.
    """
    # ── Series: convertir { '01': [{serie,correlativo_inicio}] }
    # al formato { boleta: [{serie, correlativo_inicio, activa}] }
    tipo_a_nombre = {
        "01":       "factura",
        "02":       "boleta",
        "07":       "nota_credito",
        "anulacion": "nota_debito",
    }
    series_out = {}
    for tipo_cod, lista in series_raw.items():
        nombre = tipo_a_nombre.get(tipo_cod, tipo_cod)
        items  = []
        for item in lista:
            if isinstance(item, dict):
                items.append({
                    "serie":              item.get("serie", ""),
                    "correlativo_inicio": int(item.get("correlativo_inicio", 1)),
                    "activa":             True,
                })
            else:
                items.append({"serie": str(item), "correlativo_inicio": 1, "activa": True})
        if items:
            series_out[nombre] = items

    # ── Endpoints: convertir lista dinámica del wizard
    endpoints_out = []
    nombre_sv = creds.get("nombre", "Servicio 1")
    tipo_sv   = creds.get("tipo",   "api_rest")
    usuario   = creds.get("usuario", "")
    token     = creds.get("token",   "")
    eps_lista = creds.get("endpoints", [])

    url_comp = url_anul = url_guia = url_ret = url_perc = ""
    for ep in eps_lista:
        t = ep.get("tipo", "")
        u = ep.get("url",  "").strip()
        if   t == "comprobantes": url_comp = u
        elif t == "anulaciones":  url_anul = u
        elif t == "guias":        url_guia = u
        elif t == "retenciones":  url_ret  = u
        elif t == "percepciones": url_perc = u

    endpoints_out.append({
        "nombre":           nombre_sv,
        "activo":           True,
        "formato":          "txt",
        "tipo_integracion": tipo_sv,
        "credenciales":     {"usuario": usuario, "token": token},
        "timeout":          30,
        "url_comprobantes": url_comp,
        "url_anulaciones":  url_anul,
        "url_guias":        url_guia,
        "url_retenciones":  url_ret,
        "url_percepciones": url_perc,
    })

    # ── Fuente
    tipo_fuente = fuente.get("tipo", "dbf")
    fuente_out  = {"tipo": tipo_fuente}
    if tipo_fuente in ("dbf", "excel", "csv", "access"):
        fuente_out["rutas"] = [fuente.get("ruta", "")]
    else:
        fuente_out["servidor"]   = fuente.get("host", "")
        fuente_out["base_datos"] = fuente.get("database", "")
        fuente_out["usuario"]    = fuente.get("usuario", "")
        fuente_out["puerto"]     = int(fuente.get("puerto") or
                                       {"sqlserver": 1433, "mysql": 3306,
                                        "postgresql": 5432}.get(tipo_fuente, 0))
    fuente_out["contrato_path"] = f"contratos/{cliente['cliente_id']}.yaml"

    return {
        "empresa": {
            "ruc":              cliente["ruc_emisor"],
            "razon_social":     cliente["razon_social"],
            "nombre_comercial": cliente.get("nombre_comercial", cliente["razon_social"]),
            "alias":            cliente.get("alias", ""),
            "regimen":          cliente.get("regimen", ""),
        },
        "fuente":    fuente_out,
        "series":    series_out,
        "envio":     {"endpoints": endpoints_out},
        "scheduler": {"modo": "manual", "intervalo_boletas": 10, "activo": True},
        "instalador": {"clave": "1234"},
        "licencia": {
            "clave": "",
            "endpoint_validacion": "https://licenses.disateq.com/v1/validate",
        },
    }


def _build_contrato_yaml(cliente_id: str, fuente: dict, contrato: dict) -> dict:
    """
    TASK-006 FIX: genera estructura flag_lectura/flag_escritura anidada
    que GenericAdapter._read_pending_comprobantes_dbf() espera.

    Antes (plano — roto):
        comprobantes:
          flag_campo: FLAG_ENVIO
          flag_valor: '2'

    Ahora (anidado — correcto):
        comprobantes:
          flag_lectura:  {campo: FLAG_ENVIO, valor: 2}
          flag_escritura: {campo: FLAG_ENVIO, enviado: 3, error: 4}
    """
    tipo = fuente.get("tipo", "")
    if tipo in ("dbf", "excel", "csv", "access"):
        source = {"type": tipo, "path": fuente.get("ruta", "")}
        if tipo == "dbf":
            source["encoding"] = "latin-1"
    else:
        source = {
            "type":     tipo,
            "host":     fuente.get("host", "localhost"),
            "port":     int(fuente.get("puerto") or
                           {"sqlserver": 1433, "mysql": 3306,
                            "postgresql": 5432}.get(tipo, 0)),
            "database": fuente.get("database", ""),
            "username": fuente.get("usuario", ""),
            "password": fuente.get("password", ""),
        }

    c        = contrato
    cm       = c.get("campos", {})
    ci       = c.get("items", {})
    flag_c   = c.get("flag_campo", "").strip()
    flag_v   = c.get("flag_valor", "2").strip()
    flag_t   = c.get("flag_tipo", "integer")

    # Convertir valor al tipo correcto
    try:
        flag_v_typed = int(flag_v) if flag_t == "integer" else flag_v
    except (ValueError, TypeError):
        flag_v_typed = flag_v

    # Valor enviado/error: convencion +1 / +2 respecto al pendiente
    try:
        flag_enviado = (flag_v_typed + 1) if flag_t == "integer" else "3"
        flag_error   = (flag_v_typed + 2) if flag_t == "integer" else "4"
    except TypeError:
        flag_enviado = 3
        flag_error   = 4

    comprobantes_yaml = {
        "tabla": c.get("tabla", ""),
        "flag_lectura": {
            "campo": flag_c,
            "valor": flag_v_typed,
        },
        "flag_escritura": {
            "campo":   flag_c,
            "enviado": flag_enviado,
            "error":   flag_error,
        },
    }

    # Items — solo incluir si hay datos
    items_yaml = {k: v for k, v in {
        "tabla":      ci.get("tabla", ""),
        "join_campo": ci.get("join_campo", ""),
    }.items() if v}

    # Totales — placeholder: misma tabla que comprobantes si no se especifica
    totales_yaml = {"tabla": c.get("tabla", "")}

    # Productos — placeholder vacío; el técnico lo completa en el YAML
    productos_yaml = {
        "tabla":      "",
        "join_campo": ci.get("codigo", "CODIGO_PRO"),
    }

    return {
        "cliente_id":    cliente_id,
        "source":        source,
        "comprobantes":  comprobantes_yaml,
        "items":         items_yaml,
        "totales":       totales_yaml,
        "productos":     productos_yaml,
        "cliente_varios": {
            "tipo_doc": "-",
            "num_doc":  "00000000",
            "nombre":   "CLIENTE VARIOS",
        },
    }


def _write_yaml(path: Path, data: dict) -> None:
    with open(path, "w", encoding="utf-8") as f:
        yaml.dump(data, f, allow_unicode=True,
                  default_flow_style=False, sort_keys=False)
