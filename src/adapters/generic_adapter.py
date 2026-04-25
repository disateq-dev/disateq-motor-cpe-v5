"""
generic_adapter.py
==================
Adaptador Genérico Universal — DisateQ CPE™ v4.0

Lee cualquier fuente de datos usando un contrato YAML.
Soporta: DBF, SQL Server, MySQL, PostgreSQL, XLSX, CSV.

Uso:
    from src.adapters.generic_adapter import GenericAdapter
    adapter = GenericAdapter('config/contratos/cliente_xyz.yaml')
    adapter.connect()
    pendientes = adapter.read_pending()
    for comp in pendientes:
        items = adapter.read_items(comp)
        cpe   = adapter.normalize(comp, items)
"""

import yaml
from pathlib import Path
from typing import Dict, List, Optional
from datetime import date

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))
from adapters.base_adapter import BaseAdapter


class GenericAdapter(BaseAdapter):
    """
    Adaptador genérico que usa un contrato YAML para leer cualquier fuente.
    Un solo adaptador para DBF, SQL, XLSX, CSV.
    """

    def __init__(self, contrato_path: str):
        super().__init__()
        self.contrato_path = contrato_path
        self.contrato      = self._cargar_contrato(contrato_path)
        self.fuente        = self.contrato.get('fuente', {})
        self.tipo          = self.fuente.get('tipo', '').lower()
        self._conn         = None
        self._cache_items  = {}   # cache de items por (serie, numero)
        self._cache_lookup = {}   # cache de tablas auxiliares

    # ================================================================
    # CONTRATO
    # ================================================================

    def _cargar_contrato(self, path: str) -> Dict:
        with open(path, encoding='utf-8') as f:
            c = yaml.safe_load(f)
        self._validar_contrato(c)
        return c

    def _validar_contrato(self, c: Dict):
        """Valida que el contrato tenga los campos mínimos."""
        if not c.get('fuente'):
            raise ValueError("Contrato inválido: falta sección 'fuente'")
        if not c.get('comprobantes'):
            raise ValueError("Contrato inválido: falta sección 'comprobantes'")
        campos = c['comprobantes'].get('campos', {})
        requeridos = ['tipo_doc', 'serie', 'numero']
        faltantes  = [r for r in requeridos if not campos.get(r) or '__COMPLETAR' in str(campos.get(r,''))]
        if faltantes:
            raise ValueError(f"Contrato incompleto: campos requeridos sin mapear: {faltantes}")

    # ================================================================
    # CONNECT / DISCONNECT
    # ================================================================

    def connect(self):
        if self.tipo == 'dbf':
            self._connect_dbf()
        elif self.tipo in ('sqlserver', 'mysql', 'postgres'):
            self._connect_sql()
        elif self.tipo == 'xlsx':
            self._connect_xlsx()
        elif self.tipo == 'csv':
            self._connect_csv()
        else:
            raise ValueError(f"Tipo de fuente no soportado: {self.tipo}")

    def disconnect(self):
        if self._conn:
            try:
                self._conn.close()
            except:
                pass
        self._conn = None
        self._cache_items.clear()
        self._cache_lookup.clear()

    # ================================================================
    # CONNECT DBF
    # ================================================================

    def _connect_dbf(self):
        from dbfread import DBF
        ruta = Path(self.fuente['ruta'])
        if not ruta.exists():
            raise FileNotFoundError(f"Ruta DBF no encontrada: {ruta}")

        # Cargar items en cache
        cfg_items = self.contrato.get('items', {})
        if cfg_items and cfg_items.get('tabla'):
            tabla_items = cfg_items['tabla']
            dbf_path    = ruta / f"{tabla_items}.dbf"
            if dbf_path.exists():
                self._cargar_cache_items_dbf(str(dbf_path), cfg_items)

        # Cargar tablas auxiliares (lookup)
        for nombre, cfg_lookup in self.contrato.get('lookups', {}).items():
            tabla = cfg_lookup.get('tabla')
            if tabla:
                dbf_path = ruta / f"{tabla}.dbf"
                if dbf_path.exists():
                    self._cargar_lookup_dbf(str(dbf_path), nombre, cfg_lookup)

        print(f"✅ Adaptador genérico conectado [{self.tipo.upper()}]")
        print(f"   Contrato: {Path(self.contrato_path).name}")

    def _cargar_cache_items_dbf(self, dbf_path: str, cfg: Dict):
        from dbfread import DBF
        campos    = cfg.get('campos', {})
        join_comp = cfg.get('join_campo_comprobante', '')  # campo en items que une con comprobante
        join_item = cfg.get('join_campo_items', '')        # campo en comprobante que une

        if not join_comp:
            # Intentar detectar automáticamente
            join_comp = campos.get('numero') or 'NUMERO_FAC'

        dbf = DBF(dbf_path, encoding='latin1', ignore_missing_memofile=True, raw=True)
        for record in dbf:
            try:
                dec = {k: self._decode(v) for k, v in record.items()}
                key_raw = dec.get(join_comp, '')
                if key_raw:
                    key = key_raw.strip()
                    if key not in self._cache_items:
                        self._cache_items[key] = []
                    self._cache_items[key].append(dec)
            except:
                continue

    def _cargar_lookup_dbf(self, dbf_path: str, nombre: str, cfg: Dict):
        from dbfread import DBF
        campo_clave = cfg.get('campo_clave', '')
        campo_valor = cfg.get('campo_valor', '')
        dbf = DBF(dbf_path, encoding='latin1', ignore_missing_memofile=True, raw=True)
        lookup = {}
        for record in dbf:
            try:
                dec   = {k: self._decode(v) for k, v in record.items()}
                clave = dec.get(campo_clave, '').strip()
                valor = dec.get(campo_valor, '').strip()
                if clave:
                    lookup[clave] = valor
            except:
                continue
        self._cache_lookup[nombre] = lookup

    # ================================================================
    # CONNECT SQL
    # ================================================================

    def _connect_sql(self):
        tipo      = self.tipo
        servidor  = self.fuente.get('servidor', '')
        base_datos= self.fuente.get('base_datos', '')
        usuario   = self.fuente.get('usuario', '')
        clave     = self.fuente.get('clave', '')
        puerto    = self.fuente.get('puerto')

        if tipo == 'sqlserver':
            import pyodbc
            puerto_str = f",{puerto}" if puerto else ""
            conn_str = (
                f"DRIVER={{ODBC Driver 17 for SQL Server}};"
                f"SERVER={servidor}{puerto_str};DATABASE={base_datos};"
                f"UID={usuario};PWD={clave}"
            )
            self._conn = pyodbc.connect(conn_str)

        elif tipo == 'mysql':
            import mysql.connector
            self._conn = mysql.connector.connect(
                host=servidor, database=base_datos,
                user=usuario, password=clave, port=puerto or 3306
            )

        elif tipo == 'postgres':
            import psycopg2
            self._conn = psycopg2.connect(
                host=servidor, dbname=base_datos,
                user=usuario, password=clave, port=puerto or 5432
            )

        print(f"✅ Adaptador genérico conectado [{self.tipo.upper()}]")
        print(f"   Servidor: {servidor} / BD: {base_datos}")

    # ================================================================
    # CONNECT XLSX
    # ================================================================

    def _connect_xlsx(self):
        import openpyxl
        ruta = self.fuente.get('ruta', '')
        if not Path(ruta).exists():
            raise FileNotFoundError(f"Archivo XLSX no encontrado: {ruta}")
        self._wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        print(f"✅ Adaptador genérico conectado [XLSX]")

    # ================================================================
    # CONNECT CSV
    # ================================================================

    def _connect_csv(self):
        ruta = self.fuente.get('ruta', '')
        if not Path(ruta).exists():
            raise FileNotFoundError(f"Archivo CSV no encontrado: {ruta}")
        print(f"✅ Adaptador genérico conectado [CSV]")

    # ================================================================
    # READ PENDING
    # ================================================================

    def read_pending(self) -> List[Dict]:
        cfg     = self.contrato.get('comprobantes', {})
        tabla   = cfg.get('tabla', '')
        filtro  = cfg.get('filtro', '')

        if self.tipo == 'dbf':
            return self._read_pending_dbf(tabla, filtro, cfg)
        elif self.tipo in ('sqlserver', 'mysql', 'postgres'):
            return self._read_pending_sql(tabla, filtro)
        elif self.tipo == 'xlsx':
            return self._read_pending_xlsx(tabla, filtro)
        elif self.tipo == 'csv':
            return self._read_pending_csv(filtro)
        return []

    def _read_pending_dbf(self, tabla: str, filtro: str, cfg: Dict) -> List[Dict]:
        from dbfread import DBF
        ruta     = Path(self.fuente['ruta'])
        dbf_path = ruta / f"{tabla}.dbf"
        if not dbf_path.exists():
            raise FileNotFoundError(f"Tabla no encontrada: {dbf_path}")

        # Parsear filtro simple: "FLAG_ENVIO = '2'" o "FLAG_ENVIO=2"
        filtro_campo, filtro_valor = self._parsear_filtro(filtro)

        dbf = DBF(str(dbf_path), encoding='latin1', ignore_missing_memofile=True, raw=True)
        pendientes = []
        for record in dbf:
            try:
                dec = {k: self._decode(v) for k, v in record.items()}
                if filtro_campo:
                    val = dec.get(filtro_campo, '').strip()
                    if val != filtro_valor:
                        continue
                pendientes.append(dec)
            except:
                continue

        print(f"📋 Pendientes encontrados: {len(pendientes)}")
        return pendientes

    def _read_pending_sql(self, tabla: str, filtro: str) -> List[Dict]:
        where = f"WHERE {filtro}" if filtro and '__COMPLETAR' not in filtro else ""
        cursor = self._conn.cursor()
        cursor.execute(f"SELECT * FROM [{tabla}] {where}")
        cols = [d[0] for d in cursor.description]
        rows = [dict(zip(cols, [str(v) if v is not None else '' for v in row]))
                for row in cursor.fetchall()]
        print(f"📋 Pendientes encontrados: {len(rows)}")
        return rows

    def _read_pending_xlsx(self, hoja: str, filtro: str) -> List[Dict]:
        ws   = self._wb[hoja] if hoja in self._wb.sheetnames else self._wb.active
        filas = list(ws.iter_rows(values_only=True))
        if not filas:
            return []
        encabezados = [str(c) if c else f'COL_{i}' for i, c in enumerate(filas[0])]
        filtro_campo, filtro_valor = self._parsear_filtro(filtro)
        pendientes = []
        for fila in filas[1:]:
            dec = dict(zip(encabezados, [str(v) if v is not None else '' for v in fila]))
            if filtro_campo:
                if str(dec.get(filtro_campo, '')).strip() != filtro_valor:
                    continue
            pendientes.append(dec)
        print(f"📋 Pendientes encontrados: {len(pendientes)}")
        return pendientes

    def _read_pending_csv(self, filtro: str) -> List[Dict]:
        import csv
        ruta  = self.fuente.get('ruta', '')
        sep   = self.fuente.get('separador', ',')
        enc   = self.fuente.get('encoding', 'utf-8')
        filtro_campo, filtro_valor = self._parsear_filtro(filtro)
        pendientes = []
        with open(ruta, encoding=enc, errors='replace') as f:
            reader = csv.DictReader(f, delimiter=sep)
            for row in reader:
                if filtro_campo:
                    if str(row.get(filtro_campo, '')).strip() != filtro_valor:
                        continue
                pendientes.append(dict(row))
        print(f"📋 Pendientes encontrados: {len(pendientes)}")
        return pendientes

    def _parsear_filtro(self, filtro: str):
        """Parsea filtro simple: 'CAMPO = valor' o 'CAMPO=valor'"""
        if not filtro or '__COMPLETAR' in filtro:
            return None, None
        for op in [' = ', '=', ' == ']:
            if op in filtro:
                parts = filtro.split(op, 1)
                campo = parts[0].strip()
                valor = parts[1].strip().strip("'\"")
                return campo, valor
        return None, None

    # ================================================================
    # READ ITEMS
    # ================================================================

    def read_items(self, comprobante: Dict) -> List[Dict]:
        cfg_items = self.contrato.get('items', {})
        if not cfg_items:
            return []

        cfg_comp  = self.contrato.get('comprobantes', {})
        campos_comp = cfg_comp.get('campos', {})
        numero_campo = campos_comp.get('numero', 'NUMERO_FAC')
        numero_val   = comprobante.get(numero_campo, '').strip()

        if self.tipo == 'dbf':
            return self._cache_items.get(numero_val, [])

        elif self.tipo in ('sqlserver', 'mysql', 'postgres'):
            tabla = cfg_items.get('tabla', '')
            join  = cfg_items.get('join', '')
            if join and '__COMPLETAR' not in join:
                where = f"WHERE {join.replace('comprobantes.' + numero_campo, repr(numero_val))}"
            else:
                where = ''
            cursor = self._conn.cursor()
            cursor.execute(f"SELECT * FROM [{tabla}] {where}")
            cols = [d[0] for d in cursor.description]
            return [dict(zip(cols, [str(v) if v is not None else '' for v in row]))
                    for row in cursor.fetchall()]
        return []

    # ================================================================
    # NORMALIZE
    # ================================================================

    def normalize(self, source_data: Dict, source_items: List[Dict]) -> Dict:
        cfg_comp  = self.contrato.get('comprobantes', {})
        campos    = cfg_comp.get('campos', {})
        trans     = self.contrato.get('transformaciones', {})

        # Cabecera
        tipo_raw  = self._get_campo(source_data, campos.get('tipo_doc', ''))
        tipo_map  = trans.get('tipo_doc', {})
        tipo_cpe  = tipo_map.get(tipo_raw.upper(), '03')

        serie_raw = self._get_campo(source_data, campos.get('serie', ''))
        prefijo   = {'01': 'F', '03': 'B'}.get(tipo_cpe, 'B')
        # Solo agregar prefijo si la serie no lo tiene ya
        serie     = serie_raw if serie_raw.startswith(('F','B','E')) else prefijo + serie_raw

        numero    = self._safe_int(self._get_campo(source_data, campos.get('numero', '')))
        fecha_raw = self._get_campo(source_data, campos.get('fecha', ''))
        fecha     = self._fmt_fecha(fecha_raw, trans.get('fecha', {}).get('formato_origen', 'YYYYMMDD'))

        cli_tipo  = self._get_campo(source_data, campos.get('tipo_doc_cliente', '')) or '1'
        cli_doc   = self._get_campo(source_data, campos.get('ruc_cliente', '')) or '00000000'
        cli_nom   = self._get_campo(source_data, campos.get('nombre_cliente', '')) or 'CLIENTES VARIOS'
        cli_dir   = self._get_campo(source_data, campos.get('direccion', '')) or '-'

        # Moneda
        moneda_raw = self._get_campo(source_data, campos.get('moneda', '')) or 'PEN'
        moneda_map = trans.get('moneda', {})
        moneda     = moneda_map.get(moneda_raw, moneda_raw) if moneda_raw in moneda_map else 'PEN'

        # Items
        cfg_items    = self.contrato.get('items', {})
        campos_items = cfg_items.get('campos', {})
        items_norm   = []
        total_grav = total_exon = total_igv = total_gen = 0.0

        for idx, item in enumerate(source_items, 1):
            cantidad_p = self._safe_float(self._get_campo(item, campos_items.get('cantidad', '')))
            fraccion_p = self._safe_float(self._get_campo(item, campos_items.get('cantidad_fraccion', '')))
            cantidad   = cantidad_p if cantidad_p > 0 else fraccion_p
            precio     = self._safe_float(self._get_campo(item, campos_items.get('precio', '')))
            subtotal   = self._safe_float(self._get_campo(item, campos_items.get('subtotal', '')))
            igv_item   = self._safe_float(self._get_campo(item, campos_items.get('igv', '')))
            total_item = self._safe_float(self._get_campo(item, campos_items.get('total', '')))
            codigo     = self._get_campo(item, campos_items.get('codigo', ''))
            desc_campo  = campos_items.get('descripcion', '')
            descripcion = self._get_campo(item, desc_campo)

            # Lookup descripcion desde tabla auxiliar
            if not descripcion:
                lk = self._cache_lookup.get('lookup_descripcion', {})
                if lk:
                    descripcion = lk.get(codigo.strip(), '')

            if not descripcion:
                descripcion = 'PRODUCTO SIN DESCRIPCION'

            afect = '10'  # Gravado por defecto
            total_grav += subtotal
            total_igv  += igv_item
            total_gen  += total_item

            valor_unit = round(subtotal / cantidad, 8) if cantidad else 0.0

            items_norm.append({
                'item':             idx,
                'codigo':           codigo,
                'descripcion':      descripcion or 'PRODUCTO',
                'cantidad':         cantidad,
                'unidad':           self._get_campo(item, campos_items.get('unidad', '')) or 'NIU',
                'precio_unitario':  round(total_item / cantidad, 8) if cantidad else precio,
                'valor_unitario':   valor_unit,
                'subtotal_sin_igv': round(subtotal, 2),
                'igv':              round(igv_item, 2),
                'total':            round(total_item, 2),
                'afectacion_igv':   afect,
                'unspsc':           self._get_campo(item, campos_items.get('unspsc', '')) or '10000000'
            })

        return {
            'comprobante': {
                'tipo_doc':      tipo_cpe,
                'serie':         serie,
                'numero':        numero,
                'fecha_emision': fecha,
                'moneda':        moneda
            },
            'cliente': {
                'tipo_doc':    cli_tipo,
                'numero_doc':  cli_doc,
                'denominacion':cli_nom,
                'direccion':   cli_dir
            },
            'totales': {
                'gravada':   round(total_grav, 2),
                'exonerada': round(total_exon, 2),
                'inafecta':  0.0,
                'igv':       round(total_igv,  2),
                'total':     round(total_gen,  2)
            },
            'items': items_norm
        }

    # ================================================================
    # UTILIDADES
    # ================================================================

    def _get_campo(self, record: Dict, campo: str) -> str:
        """Obtiene valor de un campo del registro, soporta campo compuesto."""
        if not campo or '__COMPLETAR' in str(campo):
            return ''
        val = record.get(campo, '')
        if val is None:
            return ''
        if isinstance(val, bytes):
            return val.decode('latin1').strip()
        return str(val).strip()

    def _decode(self, val) -> str:
        if val is None:
            return ''
        if isinstance(val, bytes):
            return val.decode('latin1').strip()
        return str(val).strip()

    def _fmt_fecha(self, val: str, formato_origen: str = 'YYYYMMDD') -> str:
        v = val.strip()
        if not v:
            return ''
        if formato_origen == 'YYYYMMDD' and len(v) == 8 and v.isdigit():
            return f"{v[6:8]}-{v[4:6]}-{v[:4]}"
        if len(v) == 10 and v[4] == '-':
            return f"{v[8:10]}-{v[5:7]}-{v[:4]}"
        if len(v) == 10 and '/' in v:
            parts = v.split('/')
            if len(parts) == 3:
                return f"{parts[0].zfill(2)}-{parts[1].zfill(2)}-{parts[2]}"
        return v

    def _safe_int(self, val) -> int:
        try:
            return int(str(val).strip().lstrip('0') or '0')
        except:
            return 0

    def _safe_float(self, val) -> float:
        try:
            return float(str(val).strip())
        except:
            return 0.0

    def get_source_info(self) -> Dict:
        return {
            'type':     'GenericAdapter',
            'fuente':   self.tipo,
            'contrato': self.contrato_path,
            'status':   'connected' if self._conn or self.tipo == 'dbf' else 'disconnected'
        }
