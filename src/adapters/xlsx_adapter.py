"""
xlsx_adapter.py
===============
Adaptador para archivos Excel (XLSX) — DisateQ POS™
"""

from typing import List, Dict
from pathlib import Path
from openpyxl import load_workbook
from .base_adapter import BaseAdapter


class XlsxAdapter(BaseAdapter):
    """Lee worksheet _CPE desde DisateQ POS™"""
    
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.workbook = None
        self.worksheet = None
        
    def connect(self) -> None:
        if not Path(self.filepath).exists():
            raise FileNotFoundError(f"Archivo no encontrado: {self.filepath}")
        
        self.workbook = load_workbook(self.filepath, data_only=True)
        
        if '_CPE' not in self.workbook.sheetnames:
            raise ValueError(f"Worksheet '_CPE' no encontrado")
        
        self.worksheet = self.workbook['_CPE']
    
    def disconnect(self) -> None:
        if self.workbook:
            self.workbook.close()
    
    def read_pending(self) -> List[Dict]:
        if not self.worksheet:
            raise RuntimeError("Debe llamar a connect() primero")
        
        comprobantes = []
        
        # Leer headers de la fila 1
        headers = {}
        for col_idx, cell in enumerate(self.worksheet[1], start=1):
            if cell.value:
                headers[col_idx] = str(cell.value).strip()
        
        # Leer datos desde fila 2
        for row_idx in range(2, self.worksheet.max_row + 1):
            row_data = {}
            tiene_datos = False
            
            for col_idx, header in headers.items():
                cell_value = self.worksheet.cell(row_idx, col_idx).value
                
                if cell_value is None:
                    cell_value = ""
                
                # Mapear nombres del POS™ a nombres internos
                mapped_header = self._map_header(header)
                row_data[mapped_header] = cell_value
                
                if cell_value not in (None, "", 0):
                    tiene_datos = True
            
            # Solo agregar si tiene datos Y tiene tipo_doc
            if tiene_datos and row_data.get('tipo_doc'):
                comprobantes.append(row_data)
        
        return comprobantes
    
    def _map_header(self, header: str) -> str:
        """Mapea nombres del POS™ a nombres internos"""
        mapping = {
            # Comprobante
            'cpe_tipo': 'tipo_doc',
            'cpe_serie': 'serie',
            'cpe_numero': 'numero',
            'cpe_fecha': 'fecha_emision',
            'cpe_moneda': 'moneda',
            
            # Cliente
            'cli_tipo_doc': 'cliente_tipo_doc',
            'cli_nro_doc': 'cliente_numero_doc',
            'cli_nombre': 'cliente_denominacion',
            'cli_direccion': 'cliente_direccion',
            
            # Los item_ ya están correctos, no necesitan mapeo
        }
        
        return mapping.get(header, header)
    
    def read_items(self, comprobante: Dict) -> List[Dict]:
        items = []
        item = {}
        
        for key, value in comprobante.items():
            if key.startswith('item_'):
                item[key] = value
        
        if item:
            items.append(item)
        
        return items
    
    def normalize(self, comp: Dict, items: List[Dict]) -> Dict:
        from decimal import Decimal
        
        tipo_doc = str(comp.get('tipo_doc', '03'))
        serie = str(comp.get('serie', 'B001'))
        numero = int(comp.get('numero', 0)) if comp.get('numero') else 0
        fecha_emision = str(comp.get('fecha_emision', ''))
        moneda = str(comp.get('moneda', 'PEN'))
        
        cliente = {
            'tipo_doc': str(comp.get('cliente_tipo_doc', '1')),
            'numero_doc': str(comp.get('cliente_numero_doc', '00000000')),
            'denominacion': str(comp.get('cliente_denominacion', 'CLIENTE VARIOS')),
            'direccion': str(comp.get('cliente_direccion', '-')),
        }
        
        totales = {
            'gravada': Decimal(str(comp.get('total_gravada', 0))),
            'exonerada': Decimal(str(comp.get('total_exonerada', 0))),
            'inafecta': Decimal(str(comp.get('total_inafecta', 0))),
            'igv': Decimal(str(comp.get('total_igv', 0))),
            'icbper': Decimal(str(comp.get('total_icbper', 0))),
            'total': Decimal(str(comp.get('total', 0))),
        }
        
        items_normalizados = []
        for item in items:
            items_normalizados.append({
                'codigo': str(item.get('item_codigo', '')),
                'descripcion': str(item.get('item_descripcion', '')),
                'unspsc': str(item.get('item_unspsc', '10000000')),
                'unidad': str(item.get('item_unidad', 'NIU')),
                'cantidad': float(item.get('item_cantidad', 0)),
                'precio_con_igv': Decimal(str(item.get('item_precio_unitario', 0))),
                'precio_sin_igv': Decimal(str(item.get('item_valor_unitario', 0))),
                'subtotal_sin_igv': Decimal(str(item.get('item_subtotal_sin_igv', 0))),
                'igv': Decimal(str(item.get('item_igv', 0))),
                'total': Decimal(str(item.get('item_total', 0))),
                'afectacion_igv': str(item.get('item_afectacion_igv', '10')),
            })
        
        # Convertir fecha
        if '-' in fecha_emision and len(fecha_emision) == 10:
            partes = fecha_emision.split('-')
            if len(partes[0]) == 4:  # YYYY-MM-DD
                fecha_str = f"{partes[2]}-{partes[1]}-{partes[0]}"
                fecha_iso = fecha_emision
            else:
                fecha_str = fecha_emision
                fecha_iso = f"{partes[2]}-{partes[1]}-{partes[0]}"
        else:
            fecha_str = fecha_emision
            fecha_iso = fecha_emision
        
        return {
            'tipo_doc': tipo_doc,
            'serie': serie,
            'numero': numero,
            'fecha_str': fecha_str,
            'fecha_iso': fecha_iso,
            'moneda': moneda,
            'cliente': cliente,
            'totales': totales,
            'items': items_normalizados,
            'forma_pago': 'Contado',
        }
